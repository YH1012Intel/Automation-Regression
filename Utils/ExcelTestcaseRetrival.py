import pandas as pd
import openpyxl

class ExcelTestcaseRetrival:
    def __init__(self, testlist_excel):
        self.testlist_excel = testlist_excel
        self.wb = openpyxl.load_workbook(self.testlist_excel)
        self.ws = self.wb.active
    
    def format_status_line(self, label, status):
        padding = max(40 - len(f"{label} [{status}]: "), 0)
        return f"{label}{' ' * padding}[{status}]"
    
    def testlist_excel_extract_excel(self):
        testcase_execute_list=[]
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            checkbox = row[0]  
            if checkbox is True:
                if row[1] == "perspec_maestro":
                    data = [row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[13]]
                elif row[1] == "pytest":
                    data = [row[1], row[2], row[3], row[10], row[11], row[12], row[13]]
                testcase_execute_list.append(data)
        return testcase_execute_list
    
    def testcase_search(self, validation_framework, item):
        #we dont know is it tag or testcase
        data = None
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if validation_framework == row[1] and item == row[3]:
                if validation_framework == "perspec_maestro":
                    data = [row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[13]]
                elif validation_framework == "pytest":
                    data = [row[1], row[2], row[3], row[10], row[11], row[12], row[13]]
        return data
    
    def feature_tag_search(self, validation_framework, item):
        #we dont know is it tag or testcase
        feature_data_list = []
        data = None
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if validation_framework == row[1] and item == row[2]:
                if validation_framework == "perspec_maestro":
                    data = [row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[13]]
                elif validation_framework == "pytest":
                    data = [row[1], row[2], row[3], row[10], row[11], row[12], row[13]]
                feature_data_list.append(data)
        return feature_data_list


    def testlist_excel_extract_argument(self, argument_testcase):
        testcase_execute_list=[]
        for validation_framework in argument_testcase:
            #print("val:", validation_framework)
            for feature_tag in argument_testcase[validation_framework]["tag"]:
                feature_data_list = self.feature_tag_search(validation_framework, feature_tag)
                if len(feature_data_list) == 0:
                    print(f"{self.format_status_line('Excel Log', 'Warning')}: Can't find testcase with {feature_tag} in excel sheet!!")
                else:
                    for testcase_data in feature_data_list:
                        if testcase_data in testcase_execute_list:
                            print(f"{self.format_status_line('Excel Log', 'Warning')}: Duplicated testcase:{testcase} in list!! Skipped append into list.")
                        else:
                            testcase_execute_list.append(testcase_data)

            for testcase in argument_testcase[validation_framework]["testcase"]:
                #print("testcase:", testcase)
                data = self.testcase_search(validation_framework, testcase)
                if data == None:
                    print(f"{self.format_status_line('Excel Log', 'Warning')}: Can't find {testcase} in excel sheet!!")
                else:
                    if data in testcase_execute_list:
                        print(f"{self.format_status_line('Excel Log', 'Warning')}: Duplicated testcase:{testcase} in list!! Skipped append into list.")
                    else:
                        testcase_execute_list.append(data)

        return testcase_execute_list


    def testlist_excel_release(self):

        df1 = pd.read_excel(self.testlist_excel)
        with pd.ExcelWriter(self.testlist_excel, engine='xlsxwriter') as writer:
            df1.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            # Add autofilter (header is at row 0 in Excel, so use row 0)
            worksheet.autofilter(0, 0, len(df1), len(df1.columns)-1)
            for row_idx in range(1, 1 + len(df1)):
                worksheet.insert_checkbox(row_idx, 0, False)
    
    def main_excel(self):
        #Extract the testcase from excel which checkbox is set to TRUE
        print(f"{self.format_status_line('Excel Log', 'info')}: Extracting testcase from excel sheet")
        testcase_execute_list = self.testlist_excel_extract_excel()
        if len(testcase_execute_list) == 0:
            print(f"{self.format_status_line('Excel Log', 'Warning')}: No checkbox is ticked in excel sheet")
        else:
            print(f"{self.format_status_line('Excel Log', 'PASS')}: Testcase successfully extracted from excel sheet")
        #Refresh the checbox to false and release the excel
        #Comment now as i dont wanna release and reset again, the function is ald proven functioning
        # print("Excel Log [info]: Refreshing checkbox in excel sheet")
        # self.testlist_excel_release()
        # print("Excel Log [PASS]: Checbox successfully unchecked in excel sheet")
        return testcase_execute_list

    def main_argument(self, argument_testcase):
        #Extract the testcase from excel base on argument
        print(f"{self.format_status_line('Excel Log', 'info')}: Extracting testcase from excel sheet")
        testcase_execute_list = self.testlist_excel_extract_argument(argument_testcase)
        print(f"{self.format_status_line('Excel Log', 'PASS')}: Testcase successfully extracted from excel sheet")
        return testcase_execute_list

# if __name__ == "__main__":
#     script_dir = os.path.dirname(os.path.abspath(__file__))
#     main_folder = os.path.dirname(script_dir)
#     testlist_excel = 'ace_testlist_ttlhg4_perspec_maestro_v1.xlsx'
#     testlist_excel = os.path.join(main_folder, testlist_excel)
    
#     function = ExcelTestcaseRetrival(testlist_excel)
#     testcase_execute_list = function.main()
#     print(testcase_execute_list)
